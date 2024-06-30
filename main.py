import openpyxl
from openpyxl import Workbook, load_workbook, worksheet
import os

from financial_title import *
from excel_function import *
from file_path import *

company_available = False
conso_available = False

STOCK = input("Please enter stock name\n").upper()
path = path + STOCK + "\\SETSmartFiles"
files = os.listdir(path)

for file in files:
    if file[-5:].upper() == ".XLSX":
        inspecting_wb = load_workbook(filename=path + "\\" + file)
        try:
            inspecting_ws = inspecting_wb["Sheet1"]
        except KeyError:
            continue
        if get_statement(inspecting_ws) == "pl" and get_type(inspecting_ws) == "Company":  # pl_company
            src_pl_comp_ws = inspecting_wb["Sheet1"]
            # dest_pl_comp_ws = inspecting_wb.copy_worksheet(src_pl_comp_ws)
        elif get_statement(inspecting_ws) == "bs" and get_type(inspecting_ws) == "Company":  # bs_company
            src_bs_comp_ws = inspecting_wb["Sheet1"]
        elif get_statement(inspecting_ws) == "cf" and get_type(inspecting_ws) == "Company":  # cf_company
            src_cf_comp_ws = inspecting_wb["Sheet1"]

        elif get_statement(inspecting_ws) == "pl" and get_type(inspecting_ws) == "Consolidate":  # pl_company
            src_pl_conso_ws = inspecting_wb["Sheet1"]
        elif get_statement(inspecting_ws) == "bs" and get_type(inspecting_ws) == "Consolidate":  # bs_company
            src_bs_conso_ws = inspecting_wb["Sheet1"]
        elif get_statement(inspecting_ws) == "cf" and get_type(inspecting_ws) == "Consolidate":  # cf_company
            src_cf_conso_ws = inspecting_wb["Sheet1"]

# check whether summarized file exist?
try:
    summarized_wb = load_workbook(path + "\\" + STOCK + "_conso.xlsx", data_only=True)
except Exception as e:
    print("sum file not found")
    summarized_wb = False
else:   # if custom field exist, store custom fields to list.
    custom_fields = []
    sum_sheet = summarized_wb["PL_conso"]
    backup_comments = back_up_comments(sheet=sum_sheet)
    sum_last_row = get_last_row(sheet=sum_sheet, col=1)
    sum_last_col = get_last_column(sheet=sum_sheet, row=1)
    first_custom_period = sum_sheet['B1'].value
    for row in range(1, sum_last_row + 1):
        if is_custom_field(sum_sheet.cell(row=row, column=1).value):    # if cell value is custom field
            custom_field = []
            for col in range(1, sum_last_col + 1):
                custom_field.append(sum_sheet.cell(row=row, column=col).value)
            custom_fields.append(custom_field)

    # save summarized to different name to back-up
    summarized_wb.save(path + "\\" + STOCK + "conso_backup.xlsx")

# if not error, create new Excel file for company financial statement
try:
    print(src_pl_comp_ws['B3'].value)
except NameError:
    print("No Company found")
else:
    company_available = True
    comp_wb = Workbook()
    pl_comp_sheet = comp_wb.create_sheet(title="PL_Company", index=1)
    bs_comp_sheet = comp_wb.create_sheet(title="BS_Company", index=2)
    cf_comp_sheet = comp_wb.create_sheet(title="CF_Company", index=3)
    copy_worksheet(src_pl_comp_ws, pl_comp_sheet)
    copy_worksheet(src_bs_comp_ws, bs_comp_sheet)
    copy_worksheet(src_cf_comp_ws, cf_comp_sheet)
    comp_wb.remove(comp_wb["Sheet"])
    # comp_wb.save("company_statement.xlsx")

# if not error, create new Excel file for consolidated financial statement
try:
    print(src_pl_conso_ws['B3'].value)
except NameError:
    print("No Consolidate found")
else:
    conso_available = True
    conso_wb = Workbook()
    pl_conso_sheet = conso_wb.create_sheet(title="PL_conso", index=1)
    bs_conso_sheet = conso_wb.create_sheet(title="BS_conso", index=2)
    cf_conso_sheet = conso_wb.create_sheet(title="CF_conso", index=3)
    copy_worksheet(src_pl_conso_ws, pl_conso_sheet)
    copy_worksheet(src_bs_conso_ws, bs_conso_sheet)
    copy_worksheet(src_cf_conso_ws, cf_conso_sheet)
    conso_wb.remove(conso_wb["Sheet"])
    # conso_wb.save("conso_statement.xlsx")

conso_sheets = []
company_sheets = []
if company_available:
    company_sheets.append(pl_comp_sheet)
    company_sheets.append(bs_comp_sheet)
    company_sheets.append(cf_comp_sheet)

if conso_available:
    conso_sheets.append(pl_conso_sheet)
    conso_sheets.append(bs_conso_sheet)
    conso_sheets.append(cf_conso_sheet)

if conso_available:
    for sheet in conso_sheets:
        # delete last 5 rows
        delete_last_5_rows(sheet)

        # delete first 12 rows
        sheet.delete_rows(1, amount=12)

        # add Period title to row 1
        sheet['A1'].value = " Period"

        # Change date format
        change_date_format(sheet)

        # Change number format for all cells
        change_number_format(sheet)

    # change cf from accumulate to quarter
    cf_accumulate_to_quarter(cf_conso_sheet)

    depreciation = get_fin_items(cf_conso_sheet, DEPRECIATION)

    oper_rev = get_fin_items(pl_conso_sheet, OPER_REV)
    cost = get_fin_items(pl_conso_sheet, COST)
    sga = get_fin_items(pl_conso_sheet, SGA)
    selling_exp = get_fin_items(pl_conso_sheet, SELLING_EXP)
    admin_exp = get_fin_items(pl_conso_sheet, ADMIN_EXP)
    total_cost_exp = get_fin_items(pl_conso_sheet, TOTAL_COST_EXP)
    net_profit = get_fin_items(pl_conso_sheet, NET_PROFIT)
    owner_net_profit = get_fin_items(pl_conso_sheet, OWN_NET_PROFIT)
    ebit = get_fin_items(pl_conso_sheet, EBIT)
    fin_cost = get_fin_items(pl_conso_sheet, FIN_COST)
    tax_exp = get_fin_items(pl_conso_sheet, TAX_EXP)
    short_debt = get_fin_items(bs_conso_sheet, SHORT_DEBT)   # Bank over draft
    short_borrowing = get_fin_items(bs_conso_sheet, SHORT_BORROWING)     # Short term borrowings
    long_debt_current = get_fin_items(bs_conso_sheet, LONG_DEBT_CURRENT)     # current portion of long term debt
    long_debt = get_fin_items(bs_conso_sheet, LONG_DEBT)

    total_short_debt = list_operation(short_debt, short_borrowing, "ADD")
    total_long_debt = list_operation(long_debt, long_debt_current, "ADD")
    total_debt = list_operation(total_short_debt, total_long_debt, "ADD")
    gross_profit = list_operation(oper_rev, cost, "SUB")
    gpm = list_operation(gross_profit, oper_rev, "DIVIDE")
    sga_to_sales = list_operation(sga, oper_rev, "DIVIDE")
    selling_exp_to_sales = list_operation(selling_exp, oper_rev, "DIVIDE")
    admin_exp_to_sales = list_operation(admin_exp, oper_rev, "DIVIDE")
    oper_profit = list_operation(gross_profit, sga, "SUB")
    ebitda = list_operation(ebit, depreciation, "ADD")
    ebitda_to_sales = list_operation(ebitda, oper_rev, "DIVIDE")
    npm = list_operation(net_profit, oper_rev, "DIVIDE")

    # npm     # below net profit
    np_row = find_row_of(NET_PROFIT, pl_conso_sheet)
    insert_list_to_excel_range(row=np_row + 1, sheet=pl_conso_sheet, items=npm, title=NPM, num_format=PERCENT_FORMAT)

    # EBITDA
    ebit_row = find_row_of(EBIT, pl_conso_sheet)
    insert_list_to_excel_range(row=ebit_row, sheet=pl_conso_sheet, items=ebitda, title=EBITDA, num_format=NUMBER)  # upper EBIT

    # %EBITDA @ lower EBITDA
    ebitda_row = find_row_of(EBITDA, pl_conso_sheet)
    insert_list_to_excel_range(row=ebitda_row + 1, sheet=pl_conso_sheet, items=ebitda_to_sales, title=EBITDA_MARGIN, num_format=PERCENT_FORMAT)

    # % admin expense to sales @ below admin
    admin_exp_row = find_row_of(ADMIN_EXP, pl_conso_sheet)
    insert_list_to_excel_range(row=admin_exp_row + 1, sheet=pl_conso_sheet, items=admin_exp_to_sales, title=ADMIN_EXP_TO_SALES, num_format=PERCENT_FORMAT)

    # % selling expense @ below selling
    selling_row = find_row_of(SELLING_EXP, pl_conso_sheet)
    if selling_row is not None:
        insert_list_to_excel_range(row=selling_row + 1, sheet=pl_conso_sheet, items=selling_exp_to_sales,
                                   title=SELLING_EXP_TO_SALES, num_format=PERCENT_FORMAT)

    # SGA to sales @ below sga
    sga_row = find_row_of(SGA, pl_conso_sheet)
    insert_list_to_excel_range(row=sga_row + 1, sheet=pl_conso_sheet, items=sga_to_sales,
                               title=SGA_TO_SALES, num_format=PERCENT_FORMAT)

    # GPM @ upper sga
    sga_row = find_row_of(SGA, pl_conso_sheet)
    insert_list_to_excel_range(row=sga_row, sheet=pl_conso_sheet, items=gpm, title=GPM, num_format=PERCENT_FORMAT)

    # Gross Profit @ above GPM
    gpm_row = find_row_of(GPM, pl_conso_sheet)
    insert_list_to_excel_range(row=sga_row, sheet=pl_conso_sheet, items=gross_profit, title=GROSS_PROFIT,
                               num_format=NUMBER)

    last_row = get_last_row(sheet=pl_conso_sheet, col=1)
    insert_list_to_excel_range(row=last_row + 1, sheet=pl_conso_sheet, items=total_debt, title=TOTAL_DEBT,
                               num_format=NUMBER)

    last_row = get_last_row(sheet=pl_conso_sheet, col=1)
    if summarized_wb:   # if summarized_wb exist
        # re-store custom fields
        for field in custom_fields:
            last_row += 1
            pl_conso_sheet.cell(row=last_row, column=1).value = field[0]
            first_custom_column = get_col_num_of(sheet=pl_conso_sheet,row=1,value=first_custom_period)
            for num, value in enumerate(field[1:]):
                pl_conso_sheet.cell(row=last_row, column=first_custom_column + num).value = value

        restore_comments(comments=backup_comments, sheet=pl_conso_sheet)

    conso_wb.save(path + "\\" + STOCK + "_conso.xlsx")

if company_available:
    for sheet in company_sheets:
        # delete last 5 rows
        delete_last_5_rows(sheet)

        # delete first 12 rows
        sheet.delete_rows(1, amount=12)

        # add Period title to row 1
        sheet['A1'].value = " Period"

        # Change date format
        change_date_format(sheet)

        # Change number format for all cells
        change_number_format(sheet)

    # change cf from accumulate to quarter
    cf_accumulate_to_quarter(cf_comp_sheet)

    depreciation = get_fin_items(cf_comp_sheet, DEPRECIATION)

    oper_rev = get_fin_items(pl_comp_sheet, OPER_REV)
    cost = get_fin_items(pl_comp_sheet, COST)
    sga = get_fin_items(pl_comp_sheet, SGA)
    selling_exp = get_fin_items(pl_comp_sheet, SELLING_EXP)
    admin_exp = get_fin_items(pl_comp_sheet, ADMIN_EXP)
    total_cost_exp = get_fin_items(pl_comp_sheet, TOTAL_COST_EXP)
    net_profit = get_fin_items(pl_comp_sheet, NET_PROFIT)
    owner_net_profit = get_fin_items(pl_comp_sheet, OWN_NET_PROFIT)
    ebit = get_fin_items(pl_comp_sheet, EBIT)
    fin_cost = get_fin_items(pl_comp_sheet, FIN_COST)
    tax_exp = get_fin_items(pl_comp_sheet, TAX_EXP)
    short_debt = get_fin_items(bs_comp_sheet, SHORT_DEBT)   # Bank over draft
    short_borrowing = get_fin_items(bs_comp_sheet, SHORT_BORROWING)     # Short term borrowings
    long_debt_current = get_fin_items(bs_comp_sheet, LONG_DEBT_CURRENT)     # current portion of long term debt
    long_debt = get_fin_items(bs_comp_sheet, LONG_DEBT)

    total_short_debt = list_operation(short_debt, short_borrowing, "ADD")
    total_long_debt = list_operation(long_debt, long_debt_current, "ADD")
    total_debt = list_operation(total_short_debt, total_long_debt, "ADD")
    gross_profit = list_operation(oper_rev, cost, "SUB")
    gpm = list_operation(gross_profit, oper_rev, "DIVIDE")
    sga_to_sales = list_operation(sga, oper_rev, "DIVIDE")
    selling_exp_to_sales = list_operation(selling_exp, oper_rev, "DIVIDE")
    admin_exp_to_sales = list_operation(admin_exp, oper_rev, "DIVIDE")
    oper_profit = list_operation(gross_profit, sga, "SUB")
    ebitda = list_operation(ebit, depreciation, "ADD")
    ebitda_to_sales = list_operation(ebitda, oper_rev, "DIVIDE")
    npm = list_operation(net_profit, oper_rev, "DIVIDE")

    # npm     # below net profit
    np_row = find_row_of(NET_PROFIT, pl_comp_sheet)
    insert_list_to_excel_range(row=np_row + 1, sheet=pl_comp_sheet, items=npm, title=NPM,
                               num_format=PERCENT_FORMAT)

    # EBITDA
    ebit_row = find_row_of(EBIT, pl_comp_sheet)
    insert_list_to_excel_range(row=ebit_row, sheet=pl_comp_sheet, items=ebitda, title=EBITDA,
                               num_format=NUMBER)  # upper EBIT

    # %EBITDA @ lower EBITDA
    ebitda_row = find_row_of(EBITDA, pl_comp_sheet)
    insert_list_to_excel_range(row=ebitda_row + 1, sheet=pl_comp_sheet, items=ebitda_to_sales, title=EBITDA_MARGIN,
                               num_format=PERCENT_FORMAT)

    # % admin expense to sales @ below admin
    admin_exp_row = find_row_of(ADMIN_EXP, pl_comp_sheet)
    insert_list_to_excel_range(row=admin_exp_row + 1, sheet=pl_comp_sheet, items=admin_exp_to_sales,
                               title=ADMIN_EXP_TO_SALES, num_format=PERCENT_FORMAT)

    # % selling expense @ below selling
    selling_row = find_row_of(SELLING_EXP, pl_comp_sheet)
    if selling_row is not None:
        insert_list_to_excel_range(row=selling_row + 1, sheet=pl_comp_sheet, items=selling_exp_to_sales,
                                   title=SELLING_EXP_TO_SALES, num_format=PERCENT_FORMAT)

    # SGA to sales @ below sga
    sga_row = find_row_of(SGA, pl_comp_sheet)
    insert_list_to_excel_range(row=sga_row + 1, sheet=pl_comp_sheet, items=sga_to_sales,
                               title=SGA_TO_SALES, num_format=PERCENT_FORMAT)

    # GPM @ upper sga
    sga_row = find_row_of(SGA, pl_comp_sheet)
    insert_list_to_excel_range(row=sga_row, sheet=pl_comp_sheet, items=gpm, title=GPM, num_format=PERCENT_FORMAT)

    # Gross Profit @ above GPM
    gpm_row = find_row_of(GPM, pl_comp_sheet)
    insert_list_to_excel_range(row=sga_row, sheet=pl_comp_sheet, items=gross_profit, title=GROSS_PROFIT,
                               num_format=NUMBER)

    last_row = get_last_row(sheet=pl_comp_sheet, col=1)
    insert_list_to_excel_range(row=last_row + 1, sheet=pl_comp_sheet, items=total_debt, title=TOTAL_DEBT,
                               num_format=NUMBER)

    comp_wb.save(path + "\\" + STOCK + "_company.xlsx")

# conso_oper_rev = get_fin_items(pl_conso_sheet, OPER_REV)
# add_fin_item(items=conso_oper_rev, sheet=pl_conso_sheet, row_to_insert=COST, position="lower", num_format=NUMBER,
#              title="OPER_REV")

# conso_wb.save("conso.xlsx")
