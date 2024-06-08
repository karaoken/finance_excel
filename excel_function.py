from openpyxl import Workbook, load_workbook, worksheet


def get_statement(sheet: worksheet):
    stmt = sheet['B4'].value
    if stmt == "Income Statement":
        return "pl"
    elif stmt == "Balance Sheet":
        return "bs"
    elif stmt == "Cash Flow":
        return 'cf'


def get_symbol(sheet: worksheet):
    return sheet['B3'].value


def get_type(sheet: worksheet):
    if sheet['B5'].value == "Company":
        return "Company"
    elif "Consolidate" in sheet['B5'].value:
        return "Consolidate"
    else:
        return None


def delete_last_5_rows(sheet: worksheet):
    last_row = get_last_row(sheet, 1)
    sheet.delete_rows(last_row - 4, amount=5)


def get_last_row(sheet: worksheet, col):
    starting_row = 500
    for row in range(starting_row, 1, -1):
        if sheet.cell(row=row, column=col).value is not None:
            return row


def get_last_column(sheet: worksheet, row):
    starting_col = 200
    for col in range(starting_col, 1, -1):
        if sheet.cell(row=row, column=col).value is not None:
            return col


def convert_date_format(date):
    if date[1:7] == "Yearly":
        year = date[10:14]
        quarter = "Q4"
    else:
        year = date[6:10]
        quarter = date[1:3]
    return f"{quarter}/{year}"


def change_date_format(sheet: worksheet):
    last_col = get_last_column(sheet, 1)
    period_row = find_row_of(" Period", sheet)
    for col in range(2, last_col + 1):
        converted_date = convert_date_format(sheet.cell(row=period_row, column=col).value)
        sheet.cell(row=period_row, column=col).value = converted_date
        # print(convert_date_format(sheet.cell(row=period_row, column=col).value))


def find_row_of(looking_title, sheet: worksheet):
    for row in range(1, 200):
        sheet_title = sheet.cell(row=row, column=1).value
        if sheet_title is None:
            return None
        if sheet_title == looking_title:
            return row
    return None


def add_gp_row(sheet: worksheet):
    sga_row = find_row_of("    Selling And Administrative Expenses", sheet)
    cost_row = find_row_of("    Costs", sheet)
    oper_rev_row = find_row_of("    Revenue From Operations", sheet)
    last_column = get_last_column(sheet, 1)

    sheet.insert_rows(sga_row, 2)
    sheet.cell(row=sga_row, column=1).value = "    Gross Profit"
    sheet.cell(row=sga_row + 1, column=1).value = "    %GPM"
    for col in range(2, last_column + 1):
        cost = sheet.cell(row=cost_row, column=col).value
        oper_rev = sheet.cell(row=oper_rev_row, column=col).value

        # Update GP row
        sheet.cell(row=sga_row, column=col).value = oper_rev - cost
        sheet.cell(row=sga_row, column=col).number_format = '#,##0.00'

        # Update GPM row
        sheet.cell(row=sga_row + 1, column=col).value = 1 - cost / oper_rev
        sheet.cell(row=sga_row + 1, column=col).number_format = "0.00%"


def add_sga_to_sales(sheet: worksheet):
    sga_row = find_row_of("    Selling And Administrative Expenses", sheet)
    oper_rev_row = find_row_of("    Revenue From Operations", sheet)
    last_col = get_last_column(sheet, 1)

    sheet.insert_rows(sga_row + 1)  # insert a row below sga row
    sheet.cell(row=sga_row + 1, column=1).value = "    % SG&A"  # add title
    for col in range(2, last_col + 1):
        sga = sheet.cell(row=sga_row, column=col).value
        oper_rev = sheet.cell(row=oper_rev_row, column=col).value
        sheet.cell(row=sga_row + 1, column=col).value = sga / oper_rev
        sheet.cell(row=sga_row + 1, column=col).number_format = '0.00%'


def change_number_format(sheet: worksheet):
    """ Change number format for a thousand Bath to a million Bath"""
    last_row = get_last_row(sheet, 1)
    last_col = get_last_column(sheet, 1)

    for row in range(2, last_row + 1):
        if "Baht/Share" in sheet.cell(row=row, column=1).value:
            continue

        for col in range(2, last_col + 1):
            value = sheet.cell(row=row, column=col).value
            if value is None or type(value) is str or value == 0:
                sheet.cell(row=row, column=col).value = ""
                continue
            sheet.cell(row=row, column=col).value = value / 1000
            sheet.cell(row=row, column=col).number_format = '#,##0.00'


def add_selling_to_sales(sheet):
    selling_exp_row = find_row_of("      Selling Expenses", sheet)
    oper_rev_row = find_row_of("    Revenue From Operations", sheet)
    last_col = get_last_column(sheet, 1)

    sheet.insert_rows(selling_exp_row + 1)  # insert a row below selling_exp row
    sheet.cell(row=selling_exp_row + 1, column=1).value = "      % Selling Expenses"
    for col in range(2, last_col + 1):
        selling_exp = sheet.cell(row=selling_exp_row, column=col).value
        oper_rev = sheet.cell(row=oper_rev_row, column=col).value
        value = selling_exp / oper_rev
        sheet.cell(row=selling_exp_row + 1, column=col).value = value
        sheet.cell(row=selling_exp_row + 1, column=col).number_format = '0.00%'


def add_admin_to_sales(sheet):
    admin_exp_row = find_row_of("      Administrative Expenses", sheet)
    oper_rev_row = find_row_of("    Revenue From Operations", sheet)
    last_col = get_last_column(sheet, 1)

    sheet.insert_rows(admin_exp_row + 1)  # insert a row below selling_exp row
    sheet.cell(row=admin_exp_row + 1, column=1).value = "      % Admin Expenses"
    for col in range(2, last_col + 1):
        admin_exp = sheet.cell(row=admin_exp_row, column=col).value
        oper_rev = sheet.cell(row=oper_rev_row, column=col).value
        value = admin_exp / oper_rev
        sheet.cell(row=admin_exp_row + 1, column=col).value = value
        sheet.cell(row=admin_exp_row + 1, column=col).number_format = '0.00%'


def add(input1, input2):
    return input1 + input2


def minus(input1, input2):
    return input1 - input2


def divide(input1, input2):
    return input1 / input2


def add_fin_item(items, sheet: worksheet, row_to_insert, position, num_format, title):
    # insert result row
    result_row = find_row_of(row_to_insert, sheet)
    if position == "lower":
        result_row += 1
    sheet.insert_rows(result_row)
    sheet.cell(row=result_row, column=1).value = title

    for index, item in enumerate(items):
        sheet.cell(row=result_row, column=2 + index).value = item


def copy_worksheet(src_sheet: worksheet, dest_sheet: worksheet):
    for row in src_sheet.iter_rows():
        for cell in row:
            dest_sheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)


def get_fin_items(sheet: worksheet, item_title) -> list:
    """
    Get financial item and return in a list.
    :param sheet:
    :param item_title: e.g. Revenue From Operation
    :return: list of financial item
    """
    result = []
    last_col = get_last_column(sheet, 1)
    last_row = get_last_row(sheet, 1)
    for row in range(1, last_row + 1):
        if sheet.cell(row=row, column=1).value == item_title:  # Found title
            for col in range(2, last_col + 1):
                result.append(sheet.cell(row=row, column=col).value)
            return result

    return result


def insert_list_to_excel_range(row, items: list, title, sheet: worksheet, num_format):
    col = 2
    sheet.insert_rows(row)
    sheet.cell(row=row, column=1).value = title

    if not items:  # if item is None or empty
        print(f"Can not assign items to worksheet: {title}")
        return

    for index, item in enumerate(items):
        sheet.cell(row=row, column=col + index).value = item
        sheet.cell(row=row, column=col + index).number_format = num_format


def cf_accumulate_to_quarter(sheet: worksheet):
    if sheet['A1'].value != " Period":
        return None

    # get last column
    last_col = get_last_column(sheet, 1)
    last_row = get_last_row(sheet, 1)
    for col in range(2, last_col):  # loop through before last column to prevent getting null for col + 1
        if int(sheet.cell(row=1, column=col).value[1]) > int(sheet.cell(row=1, column=col + 1).value[1]):
            for row in range(2, last_row + 1):
                value1 = sheet.cell(row=row, column=col).value
                value1 = 0 if value1 == '' else value1
                value2 = sheet.cell(row=row, column=col + 1).value
                value2 = 0 if value2 == '' else value2
                sheet.cell(row=row, column=col).value = value1 - value2 if value1 != value2 else ''


def list_operation(input1: list, input2: list, operation) -> list:
    result = []

    # if len(input1) != len(input2):
    #     print("inputs length are not match")
    #     return result
    if (not input1) or (not input2):  # if one of them is empty or None
        if input1:  # if input1 is not empty or None.
            return input1
        elif input2:  # if input1 is not empty or None.
            return input2
    # elif not input1:    # if empty or None
    #     input1 = [0 for _ in input2]
    # elif not input2:    # if empty or None
    #     input2 = [0 for _ in input1]

    if operation == "ADD":
        for m, n in zip(input1, input2):
            m = 0 if m == '' else m
            n = 0 if n == '' else n
            p = m + n
            if p == 0:
                result.append('')
            else:
                result.append(p)
        return result
    elif operation == "SUB":
        for m, n in zip(input1, input2):
            m = 0 if m == '' else m
            n = 0 if n == '' else n
            p = m - n
            if p == 0:
                result.append('')
            else:
                result.append(p)
        return result

    elif operation == "DIVIDE":
        for m, n in zip(input1, input2):
            m = 0 if m == '' else m
            n = 0 if n == '' else n
            if n == 0 or m == 0:
                result.append('')
            else:
                result.append(m / n)
        return result


def is_custom_field(field_name: str):
    split_field_name = field_name.split()
    if split_field_name:  # if field_name is not empty
        first_word = split_field_name[0]
        return first_word[0] == '*'
    else:
        return False


def get_col_num_of(value: str, sheet: worksheet, row: int):
    last_col = get_last_column(sheet=sheet, row=row)
    if last_col:  # if not None
        for col in range(1, last_col + 1):
            if sheet.cell(row=row, column=col).value == value:
                return col
        return None
    else:
        return None
